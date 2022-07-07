import time
import re, os
import pyautogui
from bs4 import BeautifulSoup as bs
import requests as rq
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook

class ChromeDrvier:
    @staticmethod
    def set_driver():
        # options 객체
        chrome_options = Options()

        # headless Chrome 선언
        chrome_options.add_argument('--headless')

        # 브라우저 꺼짐 방지
        chrome_options.add_experimental_option('detach', True)

        chrome_options.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.141 Whale/3.15.136.29 Safari/537.36")

        # 불필요한 에러메시지 없애기
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

        service = Service(executable_path=ChromeDriverManager().install())
        browser = webdriver.Chrome(service=service, options=chrome_options)
        browser.maximize_window()

        return browser

class AppMarketCully:
    def __init__(self):
        # ChromeDriver class에서 browser 객체 가져오기
        self.browser = ChromeDrvier().set_driver()

        # url
        self.url = 'https://www.kurly.com/shop/goods/goods_view.php?&goodsno=94594'

        # count
        self.count = 1

    def run(self)-> list:
        # Browser 이동
        self.browser.get(url=self.url)
        self.browser.implicitly_wait(5)

        # 스크롤 내리기
        self.browser.find_element(By.CSS_SELECTOR, "body").send_keys(Keys.END)
        self.browser.find_element(By.CSS_SELECTOR, "body").send_keys(Keys.PAGE_UP)
        self.browser.find_element(By.CSS_SELECTOR, "body").send_keys(Keys.PAGE_UP)

        # iframe 들어가기
        self.browser.switch_to.frame("inreview")

        return [
            self.fetch() for x in range(10) # 496
        ]

    def fetch(self)-> list:
        # 추출데이터 담을 리스트 변수
        info_list = []

        # bs로 현재 페이지소스 파싱하기
        soup = bs(self.browser.page_source, 'html.parser')
        table_lists_len = len(soup.select("div.tr_line"))

        for idx in range(table_lists_len):
            table_lists = soup.select("div.tr_line")
            table_author = table_lists[idx].select("tr > td.user_grade")[-1]

            if table_author.text != "Marketkurly":
                # 작성자 추출하기
                table_author = table_author.text.strip()

                # 리뷰 작성시간 추출하기
                table_time = table_lists[idx].select_one("tr > td.time")

                # 작성자 주문 상품명 추출하기
                review_title = table_lists[idx].select_one("div.name_purchase > p")

                # 작성자 리뷰 내용 추출하기
                review_content = table_lists[idx].select_one("div.inner_review")

                if review_title == None or review_title.text == "":
                    continue
                else:
                    review_title = review_title.text.strip()

                if table_time == None or table_time.text == "":
                    table_time = '--------'
                else:
                    table_time = table_time.text
                    table_time = re.sub('[-]', '', table_time)
                    table_time = int(table_time)

                if review_content == None or review_content.text == "":
                    continue
                else:
                    review_content = re.sub('[<div class="name_purchase">.*</div>]', '', str(review_content)).split("\n")[-2]
                    review_content = review_content.strip()

                # 추출데이터 저장
                info_list.append([self.count, review_title, table_time, table_author, review_content])

                print(
                    f"번호 : {self.count}\n작성자 주문상품명 : {review_title}\n리뷰 작성날짜 : {table_time}\n작성자 이름 : {table_author}\n상품 리뷰 내용 : {review_content}")
                print()

                # 카운트 증가
                self.count += 1

        # 페이지 이동
        next_btn = self.browser.find_element(By.CSS_SELECTOR,
                                             "a.layout-pagination-button.layout-pagination-next-page")
        self.browser.execute_script('arguments[0].click();', next_btn)

        # 추출데이터 return
        return info_list

class OpenPyXL:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(['번호', '상품옵션', '작성일', '작성자이름', '리뷰내용', ])
        self.ws.column_dimensions['A'].width = 10
        self.ws.column_dimensions['B'].width = 15
        self.ws.column_dimensions['C'].width = 25
        self.ws.column_dimensions['D'].width = 40
        self.ws.column_dimensions['E'].width = 35

        # AppMarketCully class 에서 get_content 메소드 실행 후 , return 데이터 멤버변수로 정의
        self.results : list = AppMarketCully().run()

    def savefile(self)-> None:
        row = 2
        for x  in self.results:
            for info in x:
                self.ws[f"A{row}"] = info[0]
                self.ws[f"B{row}"] = info[1]
                self.ws[f"C{row}"] = info[2]
                self.ws[f"D{row}"] = info[3]
                self.ws[f"E{row}"] = info[-1]

                row += 1

        save_Path = os.path.abspath("마켓컬리_리뷰크롤링")
        fileName = "마켓컬리_리뷰.xlsx"
        if not os.path.exists(save_Path):
            os.mkdir(save_Path)

        self.wb.save(os.path.join(save_Path, fileName))
        self.wb.close()
        pyautogui.alert(f"크롤링 데이터가 파일로 저장되었습니다\n저장된 경로는 아래와 같습니다\n\n{save_Path}")


if __name__ == "__main__":
    app = OpenPyXL()

    app.savefile()

